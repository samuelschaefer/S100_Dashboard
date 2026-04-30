"""
Created on Tue Mar  3 09:22:11 2026

@author: sschaefer
"""

import re
import time
import io
from typing import Dict, List
import streamlit as st
import numpy as np
import pandas as pd
import plotly.graph_objects as go
from dataclasses import dataclass
from openpyxl import load_workbook

color_maps = {"Azure Skies"        : ["#080708","#3772ff","#df2935","#fdca40","#e6e8e6"],
              "Heat Map"           : ["#003f5c", "#2f4b7c", "#665191", "#a05195", "#d45087", "#f95d6a", "#ff7c43","#ffa600"],
              "Divergent"          : ["#00876c", "#8eb87c", "#c2cf8c", "#f6e5a4", "#f0c07c", "#ea9960", "#e26e52", "#d43d51"],
              "Ocean Sunset"       : ["#001219","#005f73","#0a9396","#94d2bd","#e9d8a6","#ee9b00","#ca6702","#bb3e03","#ae2012","#9b2226"],
              "Fiery Sunset"       : ["#03071e","#370617","#6a040f","#9d0208","#d00000","#dc2f02","#e85d04","#f48c06","#faa307","#ffba08"],
              "Meadow Green"       : ["#d9ed92","#b5e48c","#99d98c","#76c893","#52b69a","#34a0a4","#168aad","#1a759f","#1e6091","#184e77"][::-1],
              "Dark Sunset"        : ["#335c67","#fff3b0","#e09f3e","#9e2a2b","#540b0e"],
              "Ocean Breeze"       : ["#03045e","#0077b6","#00b4d8","#90e0ef","#caf0f8"][::-1],
              "Refreshing Summmer" : ["#8ecae6","#219ebc","#023047","#ffb703","#fb8500"],
              "Watermelon"         : ["#ef476f","#ffd166","#06d6a0","#118ab2","#073b4c"][::-1],
              "Fiesta Delight"     : ["#e4572e","#17bebb","#ffc914","#2e282a","#76b041"],
              "Sunny Beach Day"    : ["#264653","#2a9d8f","#e9c46a","#f4a261","#e76f51"],
            }

_randcolorswatch = []
r = int(np.random.rand()*256)
g = int(np.random.rand()*256)
b = int(np.random.rand()*256)

for i in range(100):
    r_p = r
    g_p = g
    b_p = b
    
    while(np.sum([np.abs(r-r_p), np.abs(g-g_p), np.abs(b-b_p)]) < 150):
        r = int(np.random.rand()*256)
        g = int(np.random.rand()*256)
        b = int(np.random.rand()*256)     
    
    _randcolorswatch.append("#{0:02x}{1:02x}{2:02x}".format(r, g, b))

def clean_float(a):
    if a == None:
        return None
    try:
        return float(a)
    except ValueError:
        return np.nan

def clean_int(a):
    if a == None:
        return None
    try:
        return int(a)
    except ValueError:
        return None
#Styling Functions
def colors(a, swatch="Fiesta Delight"):
    sw =  color_maps[swatch]
    return sw[a % len(sw)]

def plcolors(a, opacity=1.0, swatch = "Heat Map"):
    return 'rgba({0},{1},{2},{3})'.format(int(colors(a, swatch)[1:3], 16),int(colors(a, swatch)[3:5], 16),int(colors(a, swatch)[5:], 16), opacity)

def randcolors(a):
    return _randcolorswatch[a % len(_randcolorswatch)]

def add_color_indicator(val, colormap):
    return f'background-color: {plcolors(val, swatch=colormap)}; color: {plcolors(val, swatch=colormap)}; border-radius: 50%'

def highlight_increase(row):
    # 'row' here only contains the columns defined in 'subset'
    styles = [''] * len(row) 
    
    for i in range(1, len(row)):
        if row.iloc[i] > row.iloc[i-1]:
            styles[i] = 'background-color: bisque'
    return styles

def color_zero(val):
    if val == 0: # Example Upper Limit
        return 'background-color: lightgreen' # Red for Fail
    return '' # Default for Pass



@dataclass
class DriverDefinition:
    driver:str
    prg_line:int
    log_alias:str
    unit:str
    down_limit:float
    up_limit:float
    binno:int
    enabled:bool

@dataclass
class TestPlanDefinition:
    program_variant:str
    device_variant:str
    name:str
    binno:int
    booking:bool
    drivers:Dict[str, DriverDefinition]  #Hold all the drivers accessed by their log alias


class ProgramDefinition:
    file_name:str
    plans:List[TestPlanDefinition] #Hold all the test plans access by their 
    device:str  #Device Variant
    program:str #Test Step, FT/QA, Temperature
    bin_definitions:Dict[int, str] = None

    def __init__(self, excel_file):  #sheet must be a openpyxl worksheet object. 
        self.plans = []
        self.file_name = excel_file.name.split('\\')[-1].split('.')[0]
        wb = load_workbook(excel_file, read_only=True, data_only=True)
        for sheet in wb.worksheets:
            if sheet.title.startswith("_"):
                if sheet.title == "_BIN_DEFINITIONS_":
                    for erow in sheet.iter_rows(values_only=True, min_row=7, min_col=1, max_col=2):
                        self.bin_definitions[int(erow[0])] = str(erow[1])
                #ignore the summary information for now.
            else:
                device_vars = {j:v.value for j,v in enumerate(sheet[2]) if v.value is not None and v.value != "DEFAULT"}
                test_step_vars = {j:v.value for j,v in enumerate(sheet[3]) if v.value is not None}
           
                for i,dev in device_vars.items():
                    for j,ts in test_step_vars.items():
                        test_plan_def = TestPlanDefinition(program_variant=ts, 
                                                           device_variant=dev, 
                                                           name=sheet.title, 
                                                           binno=clean_int(sheet.cell(row=6, column=j+1).value), 
                                                           booking=sheet.cell(row=4, column=j+1).value, 
                                                           drivers={})
                        drvs = {}
                        for i, erow in enumerate(sheet.iter_rows(values_only=True, min_row = 9)):
                            drvs[erow[3]] = DriverDefinition(driver = erow[1], 
                                                             prg_line=erow[2], 
                                                             log_alias=erow[3], 
                                                             unit=erow[4], 
                                                             down_limit=clean_float(erow[j]), 
                                                             up_limit=clean_float(erow[j+1]), 
                                                             binno=clean_int(erow[j+2]),
                                                             enabled=erow[j+3])
                        test_plan_def.drivers = drvs
                        self.plans.append(test_plan_def)                              
        wb.close()


    def listDeviceVariants(self):
        if len(self.plans) > 0:
            return list(set([plan.device_variant for plan in self.plans]))

    def listTestSteps(self):
        if len(self.plans) > 0:
            return list(set([plan.program_variant for plan in self.plans]))
        
    def getTestPlans(self, device_variant=None, program_variant=None):
        plans = self.plans
        if device_variant is not None:
            plans = [p for p in plans if p.device_variant == device_variant]
        if program_variant is not None:
            plans = [p for p in plans if p.program_variant == program_variant and p.booking ==True]
        self.device = device_variant
        self.program = program_variant
        return plans
    def getBinnoForAlias(self, log_alias:str):
        for p in self.plans:
            for k,driver in p.drivers.items():
                if k == log_alias:
                    if driver.binno:
                        return p.drivers[driver].binno
                    else:
                        return p.binno
    
    def getLogAliasForBinno(self, binno:int):  #Get a list of LOG_ALIASes for a binno
        
        if binno in [1,2,5]:
            return [] #There are no specific measurements for these bins
        else:
            alias_list = [] 
            for p in self.plans:
                for k,driver in p.drivers.items():
                    if binno == driver.binno:
                        alias_list.append(k)
            return alias_list

    def getNameForBinno(self, binno:int):  #Get the name (from _BIN_DEFINITIONS_, simplified log_alias or PLAN for a binno
        if self.bin_definitions is not None and binno in self.bin_definitions:  #Grab the bin definitions from the _BIN_DEFINTIONS_ sheet.
            return self.bin_definitions[binno]
        
        if binno == 1:
            return "GOOD"
        elif binno == 2:
            return "TIER2"   
        else:
            alias_list = [] 
            for p in self.plans:
                if p.binno is not None and p.binno == binno:
                    return f"[{p.name}]"
                for k,driver in p.drivers.items():
                    if binno == driver.binno:
                        alias_list.append(k)
            alias_list = list(set(["_".join([b for b in a.split("_") if not re.search(r"CH\d+", b)]) for a in alias_list])) #Remove all CHx from alias
            if len(alias_list) > 1:
                return "/".join(alias_list)
            elif len(alias_list) == 1:
                return alias_list[0]
            else:
                return None
    def getPlanForBinno(self, binno:int): #Get the plan for which the binno belongs to
        if binno == 1:
            return None
        else:
            plan_list = [] 
            for p in self.plans:
                if p.binno and p.binno == binno:
                    return p.name
                for k,driver in p.drivers.items():
                    if binno == driver.binno:
                        plan_list.append(p.name)
            plan_list = list(set(plan_list)) #Remove all CHx from alias
            if len(plan_list) > 1:
                return "/".join(plan_list)
            elif len(plan_list) == 1:
                return plan_list[0]
            else:
                return None
    def getBinList(self):
        return [p.binno for p in self.plans if p.binno is not None] + [dr.binno for p in self.plans for a, dr in p.drivers.items() if dr.binno is not None] 

@dataclass
class DataStructure:
    Name:str
    Data:pd.DataFrame
    Limits:pd.DataFrame
    Units:pd.DataFrame
    PrgLine:pd.DataFrame
    Metadata:dict
    TestDefinition:ProgramDefinition
    Label:str


def edit_limits(down_limit, up_limit):
    left, _, right = st.columns([2, 2, 2])
    with left:
        user_down_limit = st.number_input("Lower Limit", value=down_limit)
    with right:            
        user_up_limit = st.number_input("Upper Limit", value=up_limit)
    return user_down_limit, user_up_limit

def load_sanitize_csv(uploaded_file_stream):
    load_bar = st.progress(0.0, text="Getting metadata...")   

    uploaded_file_virtual = io.BytesIO(uploaded_file_stream)
    uploaded_file_content = uploaded_file_stream.decode("utf-8")
    log = uploaded_file_content.splitlines()
    log_info = {}

    for i,row in enumerate(log):
        cnts = row.split(',')
        if cnts[0].startswith("Program Name"):
            log_info["ProgramName"] = cnts[1].strip()
        elif cnts[0].startswith("LotID"):
            log_info["LotID"] = cnts[1].strip()
        elif cnts[0].startswith("Operator"):
            log_info["Operator"] = cnts[1].strip()
        elif cnts[0].startswith("Loadboard#"):
            log_info["Loadboard"] = cnts[1].strip()
        elif cnts[0].startswith("Handler#"):
            log_info["Handler"] = cnts[1].strip()
        elif cnts[0].startswith("Tester_ID"):
            log_info["Tester"] = cnts[1].strip()
        elif cnts[0].startswith("Test Step"):
            log_info["TestStep"] = cnts[1].strip()
        elif cnts[0].startswith("Test_Start"):
            log_info["StartDate"] = cnts[1].strip()
        elif cnts[0].startswith("Date"):
            log_info["StartDate"] = cnts[1].strip()
        elif cnts[0].startswith(".PRG Line#"):
            log_info["HeaderBlockStart"] = i
            for j,c in enumerate(cnts[1:]):
                if c.strip() != "":
                    log_info["DataCol"] = j+1
                    break
        elif cnts[0].startswith("UpLimit"):
            #up_limit = [clean_float(k.strip()) for k in cnts[start_index:]]
            pass
        elif cnts[0].startswith("DownLimit"):
            #down_limit = [clean_float(k.strip()) for k in cnts[start_index:]]
            pass
        elif cnts[0].startswith("Unit"):
            #unit_str = [c.strip() for c in cnts[start_index:]]
            #multiplier = [1.0 for c in cnts[start_index:]]
            pass
        elif cnts[0].startswith("Serial#"): 
            log_info["HeaderRow"] = i
            pass 
        elif i > 20:
            break

    uploaded_file_virtual.seek(0)    
    df_headers = pd.read_csv(uploaded_file_virtual, skiprows=log_info["HeaderBlockStart"], header=None, nrows=4)
    uploaded_file_virtual.seek(0)

    load_bar.progress(0.05, text="Loading data...")
    df_data = pd.read_csv(uploaded_file_virtual, skiprows=log_info["HeaderRow"], low_memory=False)
    df_data.columns = [ddc.strip().split('(')[0] for ddc in df_data.columns]  #Remove white space and unit
    load_bar.progress(0.45, text="Processing data...")
    
    #Sanitize the data a bit
    col_names = []
    duplicates = []
    for c in df_data.columns:
        if c not in col_names:
            col_names.append(c)
        else:
            i = 1
            while f"{c}_{i}" in col_names:
                i+=1
            col_names.append(f"{c}_{i}")
            duplicates.append(c)
    df_data.columns = col_names 
    df_headers.columns = df_data.columns
    
    df_limits = df_headers[df_headers["Serial#"].isin([c for c in df_headers["Serial#"] if "UpLimit" in c or "DownLimit" in c])].copy()
    df_units = df_headers[df_headers["Serial#"].isin([c for c in df_headers["Serial#"] if "Unit" in c])].copy()
    df_prgline = df_headers[df_headers["Serial#"].isin([c for c in df_headers["Serial#"] if "PRG Line" in c])].copy()

    for i,c in enumerate(df_data.columns):
        df_data[c] = pd.to_numeric(df_data[c], errors='coerce')
        load_bar.progress(0.5 + (i/len(df_data.columns))*0.5, text="Processing data...")
    for i,c in enumerate(df_limits.columns[1:]):
        df_limits[c] = pd.to_numeric(df_limits[c], errors='coerce')
        df_prgline[c] = pd.to_numeric(df_prgline[c], errors='coerce')
    df_units = df_units.apply(lambda x: x.str.strip() if x.dtype == "object" else x)
        

    if duplicates !=[]:
        st.warning("{0} are duplicates! I modified the names".format(", ".join(duplicates)), icon="⚠️")
    load_bar.empty()

    return log_info, df_data, {"Limits" :df_limits, "Units" : df_units, "PRGLine": df_prgline}

def select_data(df_data, df_headers, import_choice):
    print("Selecting data...")
    if import_choice == "All":
        return df_data, df_headers
    if import_choice == "Valid":
        df_data_new = df_data.dropna(axis=1, how='all')
    elif import_choice == "Condensed":
        df_data_new = df_data.dropna(axis=1, how='all')
        df_data_new = df_data_new[[col for col in df_data_new.columns if not (col.startswith("OS_") or col.startswith("ILH_") or col.startswith("ILL") or col.startswith("GPIO") or col.startswith("BIST_"))]]
    df_headers_new = {d:df[df_data_new.columns] for d,df in df_headers.items() }

    for d,df in df_headers_new.items():
        df_headers_new[d] = df.set_index("Serial#")
        df_headers_new[d].index = [i.split(':')[0].strip() for i in df_headers_new[d].index] #Clean up the spaces and colon
    return df_data_new, df_headers_new

@st.dialog("Upload Data File",  icon=":material/upload:")
def load_data_file():
    uploaded_file = st.file_uploader("Choose S100 Log File", type="csv")
    if uploaded_file:  
        # Cache the loaded data in session state to avoid reloading on every rerun
        if "current_upload" not in st.session_state or st.session_state.current_upload != uploaded_file.name:
            st.session_state.current_upload = uploaded_file.name
            st.session_state.current_log_info, st.session_state.current_df_data, st.session_state.current_df_headers = load_sanitize_csv(uploaded_file.getvalue())
            st.session_state.cached_import_choice = None  # Reset cached import choice when a new file is loaded

        log_info = st.session_state.current_log_info
        df_data = st.session_state.current_df_data
        df_headers = st.session_state.current_df_headers

        import_choice = st.radio("Select which data to import:", ["Valid", "Condensed", "All"], width="stretch", horizontal=True, help="Select between importing only valid data (removes all columns that don't have numeric values), valid data without OS, LKG, GPIO tests, or the full data set", key="import_choice_radio")        
        
        # Only call select_data() when import_choice changes
        if "cached_import_choice" not in st.session_state or st.session_state.cached_import_choice != import_choice:
            st.session_state.cached_import_choice = import_choice
            st.session_state.cached_df_data_new, st.session_state.cached_df_headers_new = select_data(df_data, df_headers, import_choice)
        
        df_data_new = st.session_state.cached_df_data_new
        df_headers_new = st.session_state.cached_df_headers_new

        label = st.text_input("File Label", value=uploaded_file.name.split('.')[0], help="Label the file data. This will be used to distinguish the data in the plots", key="file_label_input")
        if uploaded_file.name in st.session_state.file_library:
            st.badge("{0} already exist in data library!".format(uploaded_file.name), icon=":material/warning:", color="orange")
            if st.button("Replace Data", type="primary"):
                st.session_state.file_library[uploaded_file.name] = DataStructure(Name=uploaded_file.name, Data=df_data_new, Limits=df_headers_new["Limits"], Units=df_headers_new["Units"], PrgLine=df_headers_new["PRGLine"], Metadata=log_info, TestDefinition=None, Label=str(label))
                st.rerun()
        else:
            if st.button("Add Data", type="primary"):
                st.session_state.file_library[uploaded_file.name] = DataStructure(Name=uploaded_file.name, Data=df_data_new, Limits=df_headers_new["Limits"], Units=df_headers_new["Units"], PrgLine=df_headers_new["PRGLine"], Metadata=log_info, TestDefinition=None, Label=str(label))
                st.rerun()

def generate_traces(series_list:list, file_plot_options:str, site_plot_options):
    traces = {}
    limits = {}
    unique_files = list(set([s[0] for s in series_list]))
    unique_measurements = list(set([s[1] for s in series_list]))
    if file_plot_options == "Combine All":
        for m in unique_measurements:
           traces[m] = pd.concat([lib[f].Data[["Serial#", "Site#", m]] for f in lib], axis=0)
           limits[m] = (np.max(np.array([lib[p[0]].Limits.loc["DownLimit", p[1]] for p in plot_series if p[1] == m])), 
                   np.min(np.array([lib[p[0]].Limits.loc["UpLimit", p[1]] for p in plot_series if p[1] == m])))
    elif file_plot_options == "Compare All" or file_plot_options == "Select...":
        for m in series_list:
            series_name = f"{m[1]} [{lib[m[0]].Label}]"
            df = lib[m[0]].Data[["Serial#", "Site#", m[1]]].rename(columns={m[1]: series_name,})
            traces[series_name] = df
            limits[series_name] = (lib[m[0]].Limits.loc["DownLimit", m[1]],lib[m[0]].Limits.loc["UpLimit", m[1]])
    else:
        if len(unique_files) != 1:
            raise Exception("Code error. Why is there a single file selected, but more than one file listed in the plot_series?")
        for m in unique_measurements:
            traces[m] = lib[unique_files[0]].Data[["Serial#", "Site#", m]]
            limits[m] =  (lib[unique_files[0]].Limits.loc["DownLimit", m],lib[unique_files[0]].Limits.loc["UpLimit", m])

    if site_plot_options == "Compare":
        traces_site = {}
        limits_site = {}
        for name,df in traces.items():
            for site in df["Site#"].unique():
                series_name = f"{name} [S{site}]"
                traces_site[series_name] = df[df['Site#'].isin([site])].rename(columns={name: series_name,})
                limits_site[series_name] = limits[name]
        return traces_site, limits_site
    elif site_plot_options.isnumeric():
        for name,df in traces.items():
            traces[f"{name}"] = df[df['Site#'].isin([int(site_plot_options)])]

    return traces, limits

@st.dialog("Upload Excel Limits Document",  icon=":material/discover_tune:")
def load_test_definition(lib_key):
    uploaded_definition_file = st.file_uploader("Choose Test Definition Excel File for this lot", type="xlsx")
    
    if uploaded_definition_file:
        test_def = ProgramDefinition(uploaded_definition_file)
        #Display three controls, a drop down to select device variant, a drop down to select test step and a button to close everything and copy it. 
        apply_all_data = st.checkbox("Apply to all datasets", value=False)
        dev_dd, teststep_dd, finish_bt  = st.columns(3, vertical_alignment="bottom", gap="medium")
        with dev_dd:
            device_variant = st.selectbox("Device Variant", options=test_def.listDeviceVariants(), help="Select the device variant for this test definition.")
        with teststep_dd:
            test_step = st.selectbox("Test Step", options=test_def.listTestSteps(), help="Select the test step for this test definition.")
            
        with finish_bt:
            if st.button("Done", type="primary", width="stretch"):
                test_def.plans = test_def.getTestPlans(device_variant, test_step)
                if apply_all_data:
                    for lk in st.session_state.file_library:
                        st.session_state.file_library[lk].TestDefinition = test_def
                else:
                    st.session_state.file_library[lib_key].TestDefinition = test_def
                st.rerun()

def styled_box(text, bg_color="#E3F2FD", border_color="#2196F3"):
    return f"""
    <div style="
        background-color: {bg_color};
        border: 1px solid {border_color};
        color: {border_color};
        padding: 2px 10px;
        text-align: center;
        border-radius: 8px;
        font-weight: bold;
        font-size: 12px;
        margin-bottom: 2px;
        margin-right:20px;
        margin-left:20px;
    ">
        {text}
    </div>
    """

@st.dialog("Bin Details",  icon=":material/read_more:", width="large"   )
def show_bin_details(selected_row, color_map):
    for c in ["Bin Code", "Alias", "Plan"]:
        st.write(f"{c} : \t\t{selected_row.loc[selected_row.index[0], c]}")
    
    bin_code = selected_row.loc[selected_row.index[0], "Bin Code"]
    current_plans = [text for text in selected_row.loc[selected_row.index[0], "Plan"].split('/')]
    current_aliases = {f: ds.TestDefinition.getLogAliasForBinno(bin_code) for f,ds in st.session_state.file_library.items()}

    fixed_height=800

    test_def_variants = list(set([f"{ds.TestDefinition.file_name} : {ds.TestDefinition.device},{ds.TestDefinition.program}" for ds in st.session_state.file_library.values()]))
    if len(test_def_variants) > 1:
        select_def = st.selectbox("Select the test definition you want to view:", options=test_def_variants)
    else:
        select_def = test_def_variants[0]

    for ds in  st.session_state.file_library.values():
        if ds.TestDefinition.file_name == select_def.split(':')[0].strip() and \
        ds.TestDefinition.device == select_def.split(':')[1].strip().split(',')[0].strip() and \
        ds.TestDefinition.program == ",".join(select_def.split(':')[1].strip().split(',')[1:]):
            test_definition = ds.TestDefinition
    dist, flow = st.columns([6,4])
    with dist:
        
        #traces returns a dict of dataframes that include a shallow copy of the source dataframe in lib[]. 
        #the dataframe includes the renamed series (including file/site specifics) and a couple metadata columns (Serial#, Site#)
        traces = {}
        limits = {}
        units = []
        for f, aliases in current_aliases.items():
            ds = st.session_state.file_library[f]
            for m in aliases:
                if m in ds.Data.columns:
                    series_name = f"{m} [{ds.Label}]"
                    traces[series_name] = ds.Data[["Serial#", m]].rename(columns={m: series_name,})
                    limits[series_name] = (ds.Limits.loc["DownLimit", m],ds.Limits.loc["UpLimit", m])
                    units.append(ds.Units.loc["Unit",m])
        if traces == {}:
            st.info("This measurement has no data to plot.")
        else:
            bins = 80 
            x_max_limit = max([l[1] for l in limits.values()])
            x_min_limit = min([l[0] for l in limits.values()])
            bin_size = (x_max_limit - x_min_limit) / bins 
            x_bin_start = np.min([np.min(np.array([d[n].min() for n,d in traces.items()]))-bin_size/2.0, x_min_limit])

            #Histogram
            fig_dist_dg = go.Figure()  
            for i, (name,df) in enumerate(traces.items()):          
                fig_dist_dg.add_trace(go.Histogram(
                    x=df[name],
                    name=name.replace('[', "<b>[").replace("]", "]</b>"),
                    marker=dict(
                        color=plcolors(i, 0.5, cmap),
                        line=dict(
                            color=plcolors(i, 1.0, cmap),
                            width=1)
                        ),
                    xbins= dict(size=bin_size, start=x_bin_start)
                ))
            
            fig_dist_dg.update_layout(
                #title_text='Measurement Distribution',
                height = 600,
                xaxis_title_text=f"[{",".join(list(set(units)))}]",
                xaxis=dict(range=[x_min_limit, x_max_limit]),
                yaxis_title_text='device count',
                barmode="overlay",
                bargap=0.0, # The gap between bars
                legend=dict(orientation="h", # Horizontal legend (like a footer)
                            yanchor="top",
                            y=-0.15,
                            xanchor="center",
                            x=0.5)
            ) 
            with st.container(height=fixed_height, border=False):
                st.subheader("Distribution", text_alignment="center")
                st.plotly_chart(fig_dist_dg, use_container_width=True)

    with flow:
        arrow_html = """
        <div style="text-align: center; color: #757575; font-size: 12px; margin: -5px 0;">
            &darr;
        </div>
        """
        with st.container(height=fixed_height, border=False):
            st.subheader("Program Flow", text_alignment="center")
            # Apply custom CSS to the container to ensure centering
            st.markdown(
                    """
                    <style>
                    .flow-container {
                        display: flex;
                        flex-direction: column;
                        align-items: center;
                        width: 300px; /* or auto */
                        margin: auto;
                    }
                    </style>
                    """,
                    unsafe_allow_html=True
                )
            st.markdown('<div class="flow-container">', unsafe_allow_html=True)
            

            for i,plan in enumerate(test_definition.plans):
                    
                st.markdown(styled_box(plan.name,  plcolors(1 if plan.name in current_plans else 0, 0.5, color_map), plcolors(1 if plan.name in current_plans else 0, 1.0, color_map)), unsafe_allow_html=True)
                if i < len(test_definition.plans) - 1:
                    st.markdown(arrow_html, unsafe_allow_html=True)

            st.markdown('</div>', unsafe_allow_html=True)

                

#Page Config
st.set_page_config(page_title="Production Visualizer", layout="wide")

st.markdown("""
    <style>
    /* Remove padding from the main block */
    .block-container {
        padding-top: 4rem;
        padding-bottom: 0rem;
        padding-left: 1rem;
        padding-right: 1rem;
    }

     /* Optional: Reduce space between elements */
    [data-testid="stVerticalBlock"] > div {
       gap: 0.5rem;
    }
    /* Target only tables inside a specific div class */
    .compact-table table {
        margin-bottom: 0px !important;
    }
    .compact-table td, .compact-table th {
        padding: 2px 8px !important; /* Extremely tight padding */
        line-height: 1.0 !important;
        font-size: 13px !important;
    }
    /* Sync the button alignment */
    .stButton > button {
        margin-top: -5px; /* Pull buttons closer to the table */
        width: 100%;
    }
    </style>
    """, unsafe_allow_html=True)


# session_state variables
if 'file_library' not in st.session_state:
    st.session_state.file_library = {}
lib: dict[str, DataStructure] = st.session_state.file_library

# Sidebar
if st.session_state.file_library != {}:    
    #Sidebar
    st.sidebar.header("Yield Options")
    yd_from_binno = st.sidebar.radio("Yield Source:", ["Bin#", "Raw Data"], horizontal=True, help="Select whether to calculate yield bins from the data or from the bin results. If using Raw Data, only those bins that have been loaded will be shown.")
    yd_plot_bin_1 = st.sidebar.toggle("Show Good Bins", value=True, help="Toggle whether to include the GOOD/TIER2/PREMIUM bins in the yield summary.")
    yd_bar_count = st.sidebar.select_slider("Yield Bin Count:", list(range(7,17)) + ['All'], value=7, help="Select the number of bars to show on the yield summary.")

    st.sidebar.header("Plot Options")
    file_to_plot = st.sidebar.selectbox("Source File", [ds.Label for ds in lib.values()] + (["Combine All", "Compare All", "Select..."] if len(lib) > 1 else []), label_visibility="visible", index=0, help="Select the file for which to plot data. ")
    
    plot_series = []
    if file_to_plot == "Select...": #Add any traces from any of the files. Allows users to plot measurements with different labels. 
        for file in lib:
            dataCol = lib[file].Metadata["DataCol"]
            selected_series = st.sidebar.multiselect(
                 f"{lib[file].Label} Measurements",
                 options=lib[file].Data.columns.to_list()[dataCol:],
            )
            plot_series += [(file, ss) for ss in selected_series] 

    elif file_to_plot == "Compare All" or file_to_plot == "Combine All":  #We only list columns that are present in ALL loaded files           
        select_options = []
        for file in lib:
            dataCol = lib[file].Metadata["DataCol"]
            select_options += [o for o in lib[file].Data.columns.to_list()[dataCol:] if o not in select_options]
        selected_series = st.sidebar.multiselect(
                f"All Measurements",
                options=select_options,
            )
        for file in lib:
            plot_series += [(file, ss) for ss in selected_series] 
    else: 
        file = [f for f in lib if lib[f].Label == file_to_plot][0]
        dataCol = lib[file].Metadata["DataCol"]
        selected_series = st.sidebar.multiselect(
                f"{lib[file].Label} Measurements",
                options=lib[file].Data.columns.to_list()[dataCol:],
            )
        plot_series += [(file, ss) for ss in selected_series] 
    
    plot_library =[f for f in list(set([pd[0] for pd in plot_series]))]
    test_sites = list(set(str(item) for sublist in  [lib[file].Data["Site#"].unique() for file in plot_library] for item in sublist))
    if len(test_sites) > 0:
        if len(test_sites) > 1:
            test_sites += ["Both", "Compare"]
            test_site_default = "Both"
        else:
            test_site_default = test_sites[0]
        combine_sites = st.sidebar.segmented_control("Site Options", test_sites, selection_mode="single", default=test_site_default, label_visibility="visible", width="stretch")
        if combine_sites == None:
            combine_sites = test_site_default

    cmap = st.sidebar.selectbox("Color Theme", [cm for cm in color_maps], index=0, label_visibility="visible", accept_new_options=False, width="stretch")

    show_full_x_range = st.sidebar.toggle("Plot Full Range", value=False, help="Toggle whether to show the full x range of the data or just the area around the limits. This can help visualize how much of the data is clustered around the limits.")
    subsample_data = st.sidebar.toggle("Subsample Data", value=True if max([len(ds.Data) for ds in lib.values()]) > 5000 else False, help="Toggle whether to subsample the data when plotting. This can help with performance when plotting large datasets, but may make the plots less accurate.")
    #draw_limits = st.sidebar.toggle("Toggle Limits")

# Data Library
with st.expander("Data Library", expanded=True, icon=":material/library_books:"):
    if st.session_state.file_library == {}:
        st.info("Load a production log .csv file to get started.")
        if st.button("Load Data", icon=":material/upload:", width="stretch"):
            load_data_file()
    else:
        num_cols = len(lib) + 1 if len(lib) < 4 else len(lib)
        #st.header("Data Library")
        df_files_data=[
            [ds.Metadata["LotID"] for ds in lib.values()],
            [len(ds.Data) for ds in lib.values()],
            ["{0:.2f}%".format((np.count_nonzero(ds.Data["Bin#"] == 1) + np.count_nonzero(ds.Data["Bin#"] == 2) + np.count_nonzero(ds.Data["Bin#"] == 5)) / len(ds.Data) * 100.0) for ds in lib.values()],
            [ds.Metadata["StartDate"] for ds in lib.values()],
            [f"{ds.TestDefinition.file_name} : {ds.TestDefinition.device}, {ds.TestDefinition.program}" if ds.TestDefinition else "" for ds in lib.values()],
            [ds.Label for ds in lib.values()],
            [ds.Metadata["TestStep"] for ds in lib.values()],
            [ds.Metadata["Handler"] for ds in lib.values()],
            [ds.Metadata["Tester"] for ds in lib.values()],
            ["{0} ({1})".format("Single" if ds.Data["Site#"].unique().sum() == 1 else "Dual", ",".join([str(s) for s in ds.Data["Site#"].unique()])) for ds in lib.values()],
            [ds.Metadata["ProgramName"] for ds in lib.values()],
        ]

        df_files = pd.DataFrame(df_files_data, index=["Lot", "Device Count", "Yield", "Test Time", "Program", "Label", "Test Step", "Handler", "Tester", "Sites", "PGM Location"], columns=[ds.Name for ds in lib.values()])
        
        data_files_cols = st.columns([0.9 / (num_cols-1) for i in range(num_cols-1)] + [0.1], gap="xsmall", vertical_alignment="center") if len(lib) < 4 else st.columns(num_cols, gap="xsmall", vertical_alignment="center")
    
        for i, col_name in enumerate(df_files.columns):
            with data_files_cols[i]:
                if st.session_state.get(f"remove_{col_name}"):
                    lib.pop(col_name)
                    st.rerun()

                st.dataframe(df_files[col_name], height="content", row_height=30)  
                if lib[col_name].TestDefinition is None:
                    if st.button("Load Test Definition", icon=":material/discover_tune:", width="stretch", key=f"load_testdef_{col_name}"):
                        load_test_definition(col_name)


                st.button(f"Remove {lib[col_name].Label}", icon=":material/delete_forever:", width="stretch", key=f"remove_{col_name}")
        
        if num_cols > len(df_files.columns):
            with data_files_cols[-1]:
                if st.button("Add", icon=":material/add:", width="stretch", type="primary"):
                    load_data_file()

# Yield
if st.session_state.file_library != {}:
    with st.expander("Yield", expanded=True, icon=":material/paid:"):
        num_cols = len(lib) + 1 if len(lib) < 4 else len(lib)
        data_files_cols = st.columns([0.9 / (num_cols-1) for i in range(num_cols-1)] + [0.1], gap="xsmall") if len(lib) < 4 else st.columns(num_cols, gap="xsmall")
        fig_yields = [go.Figure() for i in range(len(lib))]
        for i,(name,ds) in enumerate(lib.items()):
            #Get yield data for this file
           
            yield_values = []
            
            if yd_from_binno == "Raw Data" :
                for col in ds.Data.columns[ds.Metadata["DataCol"]:]:
                    failures_mask = (ds.Data[col] > ds.Limits.loc["UpLimit", col]) | (ds.Data[col] < ds.Limits.loc["DownLimit", col])
                    col_idx = ds.Data.columns.get_loc(col)
                    next_col_data = ds.Data.iloc[failures_mask.values, col_idx+1] if col_idx+1 < len(ds.Data.columns) else None
                    if failures_mask.sum() > 0 and (next_col_data is None or (next_col_data is not None and next_col_data.isna().any())) and 'READ_S2M_BINARY' not in col:  #If there are failures, and the next column is empty (which is where bin codes are logged by default) and this column is not a REG_READ value (which are not logged correctly by default), then we can get bin code information for these failures.
                        bin_codes = ds.Data.loc[failures_mask, "Bin#"].unique()  
                        if 1 in bin_codes or 2 in bin_codes or 5 in bin_codes:
                            print(f"Unexpected GOOD/TIER2/PREMIUM bin codes found for failures in column {col} of file {ds.Label}.")                 
                        yield_values.append([failures_mask.sum()/ len(ds.Data) * 100.0, ",".join([str(b) for b in bin_codes])])
                    else:
                        bin_codes = []
                        yield_values.append([0, ",".join([str(b) for b in bin_codes])])

                if yd_plot_bin_1:           
                    yield_values.append([np.count_nonzero(ds.Data["Bin#"] == 1) / len(ds.Data) * 100.0, 1])
                    yield_values.append([np.count_nonzero(ds.Data["Bin#"] == 2) / len(ds.Data) * 100.0, 2])
                    yield_values.append([np.count_nonzero(ds.Data["Bin#"] == 5) / len(ds.Data) * 100.0, 5])
                    yield_df = pd.DataFrame(yield_values, index=pd.Index(list(ds.Data.columns[ds.Metadata["DataCol"]:]) + ["GOOD", "TIER2", "PREMIUM"]), columns=["Yield Loss", "Bin Codes"])
                else:
                    yield_df = pd.DataFrame(yield_values, index=ds.Data.columns[ds.Metadata["DataCol"]:] , columns=["Yield Loss", "Bin Codes"])
                #yield_df_plot = yield_df[yield_df["Yield Loss"] > 0.5]
                missing_bins_loss = 100.0 - sum([b[0] for b in yield_values]) if yd_plot_bin_1 else 100.0 - sum([b[0] for b in yield_values]) - ((np.count_nonzero(ds.Data["Bin#"] == 1) + np.count_nonzero(ds.Data["Bin#"] == 2) + np.count_nonzero(ds.Data["Bin#"] == 5)) / len(ds.Data) * 100.0)

            else:
                for bin_code in ds.Data["Bin#"].unique():
                    bin_mask = ds.Data["Bin#"] == bin_code
                    if bin_code not in [1,2,5] or yd_plot_bin_1:
                        yield_values.append([bin_mask.sum() / len(ds.Data) * 100.0, bin_code])
                yield_df = pd.DataFrame(yield_values, index=[f"{ds.TestDefinition.getNameForBinno(b[1])}" for b in yield_values] if ds.TestDefinition is not None else [f"Bin# {b[1]}" for b in yield_values], columns=["Yield Loss", "Bin Codes"])
            
                
            yield_df_plot = yield_df.sort_values("Yield Loss", ascending=False)
            if yd_bar_count == "All":
                yield_df_plot = yield_df_plot[yield_df_plot["Yield Loss"] > 0.0]
            else:
                yield_df_plot = yield_df_plot.head(int(yd_bar_count))
            
            missing_bins_loss = 100.0 - sum([b[0] for b in yield_values]) if yd_plot_bin_1 else 100.0 - sum([b[0] for b in yield_values]) - ((np.count_nonzero(ds.Data["Bin#"] == 1) + np.count_nonzero(ds.Data["Bin#"] == 2) + np.count_nonzero(ds.Data["Bin#"] == 5)) / len(ds.Data) * 100.0)
            if missing_bins_loss > 0.0001:
                yield_df_plot = pd.concat([yield_df_plot, pd.DataFrame([[missing_bins_loss, "UNKNOWN/MISSING"]], index=["UNKNOWN/MISSING"], columns=["Yield Loss", "Bin Codes"])])
            with data_files_cols[i]:   
                fig_yields[i].add_trace(go.Bar(
                    x = yield_df_plot["Yield Loss"],
                    y = [f"{name} <b>({ yield_df_plot.loc[name,"Bin Codes"]})</b>" for name in yield_df_plot.index],
                    text = ["{0:.2f}%".format(yl) for yl in yield_df_plot["Yield Loss"]],
                    orientation = "h",
                    textposition="auto", 
                    marker=dict(
                        color=plcolors(i, 0.5, cmap),
                        line=dict(
                            color=plcolors(i, 1.0, cmap),
                            width=1)
                        ),
                ))
                fig_yields[i].update_layout(
                        title = f"{ds.Label} Yield Summary",
                        yaxis=dict(autorange="reversed"),
                        bargap=0.2,
                    )
                st.plotly_chart(fig_yields[i])

# Distribution Plots
    with st.expander("Plots", expanded=True, icon=":material/bar_chart:"):
        if plot_series != []:
            
            #traces returns a dict of dataframes that include a shallow copy of the source dataframe in lib[]. 
            #the dataframe includes the renamed series (including file/site specifics) and a couple metadata columns (Serial#, Site#)

            traces, limits = generate_traces(plot_series, file_to_plot, combine_sites)

            x_max_limit = np.max(np.array([lib[p[0]].Limits.loc["UpLimit", p[1]] for p in plot_series]))
            x_min_limit = np.min(np.array([lib[p[0]].Limits.loc["DownLimit", p[1]] for p in plot_series]))
            
            bins = 80 
            bin_size = (x_max_limit - x_min_limit) / bins 

            units = [lib[p[0]].Units.loc["Unit", p[1]] for p in plot_series]

            #Create a list of all the traces ("legend name", "series", )
              
            x_max_data = np.max(np.array([d[n].max() for n,d in traces.items()]))
            x_min_data = np.min(np.array([d[n].min() for n,d in traces.items()]))   
            x_bin_start = np.min([np.min(np.array([d[n].min() for n,d in traces.items()]))-bin_size/2.0, x_min_limit])           
            

            #Histogram
            fig_dist = go.Figure()  
            for i, (name,df) in enumerate(traces.items()):          
                fig_dist.add_trace(go.Histogram(
                    x=df[name] if not subsample_data else df[name].sample(frac=0.1, random_state=1),
                    name=name.replace('[', "<b>[").replace("]", "]</b>"),
                    marker=dict(
                        color=plcolors(i, 0.5, cmap),
                        line=dict(
                            color=plcolors(i, 1.0, cmap),
                            width=1)
                        ),
                    xbins= dict(size=bin_size, start=x_bin_start)
                ))
            
            fig_dist.update_layout(
                title_text='Measurement Distribution',
                height = 600,
                xaxis_title_text=f"[{",".join(list(set(units)))}]",
                xaxis=dict(range=[x_min_data if show_full_x_range else x_min_limit, x_max_data if show_full_x_range else x_max_limit]),
                yaxis_title_text='device count',
                barmode="overlay",
                bargap=0.0, # The gap between bars
                legend=dict(orientation="h", # Horizontal legend (like a footer)
                            yanchor="top",
                            y=-0.15,
                            xanchor="center",
                            x=0.5)
            ) 
            
            fig_trends = go.Figure()
            for i, (name,df) in enumerate(traces.items()):
                fig_trends.add_trace(go.Scattergl(
                    x=df['Serial#'] if not subsample_data else df['Serial#'][::10], 
                    y=df[name] if not subsample_data else df[name][::10], 
                    mode='markers', 
                    marker=dict(
                            size=1,
                            color=plcolors(i, 1.0, cmap),
                            line=dict(
                                width=2,
                                color=plcolors(i, 1.0, cmap),
                            )
                        ),
                    name=name.replace('[', "<b>[").replace("]", "]</b>"),
                ))

            fig_trends.update_layout(title="Measurement Trend", 
                        height = 600,
                        xaxis_title="serial #", 
                        yaxis_title=f"[{",".join(list(set(units)))}]",
                        showlegend=True,
                        legend=dict(orientation="h", # Horizontal legend (like a footer)
                            yanchor="top",
                            y=-0.15,
                            xanchor="center",
                            x=0.5)
                        )
            col1, col2 = st.columns(2)
            with col1:
                st.plotly_chart(fig_dist, use_container_width=True, config={"toImageButtonOptions": {"format": "png"}})
            with col2:
                st.plotly_chart(fig_trends, use_container_width=True, config={"toImageButtonOptions": {"format": "png"}})

            #Calculate summary statistics
            summary_table = []
            for i,(n,limit) in enumerate(limits.items()):
                summary_table.append({
                    "C" : i,
                    "Trace" : n, 
                    "Mean" : traces[n][n].mean(),
                    "Max"  : traces[n][n].max(),
                    "Min" : traces[n][n].min(),
                    "Std Dev" : traces[n][n].std(),
                    "Count" : traces[n][n].count(),
                    "Down Limit" : limit[0],
                    "Up Limit" : limit[1],
                    "Fail Count" : ((traces[n][n] < limit[0]) | (traces[n][n] > limit[1])).sum(),
                    "Yield (%)" : ((traces[n][n] >= limit[0]) & (traces[n][n] <= limit[1])).sum()  /  traces[n][n].count() * 100,
                })
            df_summary = pd.DataFrame(summary_table).set_index("C")            
            styled_df = df_summary.reset_index().style.map(add_color_indicator, subset=['C'], colormap=cmap)

        else:
            st.info("Select a result to start plotting data.")

# Statistics
    with st.expander("Statistics", expanded=True, icon=":material/analytics:"):
        if plot_series != []:
            dec_count = int(max([np.ceil(np.max(-1.0* np.log10(df_summary['Std Dev'].to_numpy()))), np.ceil(np.max(-1.0* np.log10(df_summary['Mean'].to_numpy()))), 2]))
            st.dataframe(
                styled_df,
                column_config={
                    "C": st.column_config.TextColumn(" ", width=20, help="Trace Hex Code"),
                    "Mean": st.column_config.NumberColumn("Mean", format="%.{0}f".format(dec_count)),
                    "Min": st.column_config.NumberColumn("Min", format="%.{0}f".format(dec_count-1)),
                    "Max": st.column_config.NumberColumn("Max", format="%.{0}f".format(dec_count-1)),
                    "Std Dev": st.column_config.NumberColumn("Std Dev", format="%.{0}f".format(dec_count+1)),
                    "Down Limit": st.column_config.NumberColumn("Down Limit", format="%.{0}f".format(dec_count-1)),
                    "Up Limit": st.column_config.NumberColumn("Up Limit", format="%.{0}f".format(dec_count-1)),
                    "Yield (%)": st.column_config.NumberColumn("Yield (%)", format="%.2f")  
                },
                hide_index=True,
                row_height=30,
                #use_container_width=False,
                
            )
            
        else:
            st.info("Select a result to start viewing statistics")
#Specific Analysis
    with st.expander("Advanced Analysis", expanded=True, icon=":material/labs:"):
        eng_task_list = ["1 - Detailed Yield analysis", 
                         "2 - Detailed Bin Analysis"]  #Add more as needed
        eng_task_list_description = ["Determine yield progressions across test steps to see what parameters are sensitive to re-testing and how it affects yield.",
                                     "List all bins the number of failing devices for each of them. Use this data to eliminate potential test items with no failures.", 
                                     ]

        eng_task = st.selectbox("Select Analysis Task",
                                eng_task_list,
                                placeholder="Select Engineering Task",
                                index=None, 
                                label_visibility="collapsed")
        #st.info(eng_task_list_description[eng_task_list.index(eng_task)] if eng_task is not None else "Select an egineering task to get started")

        if eng_task is None:
            pass 
        elif eng_task.startswith("1 "):
            st.warning("Not implemented yet. Please check back later")
        elif eng_task.startswith("2 "):
            if ds.TestDefinition == None:
                st.error(f"You must link a test definition (Master Limits) to the dataset(s) to check for unused bins.")
            else:
                if len(list(set([f"{ds.TestDefinition.file_name}:{ds.TestDefinition.device},{ds.TestDefinition.program}" for ds in lib.values()]))) > 1:
                    st.warning("It looks like the different datasets have unique test definitions. If the bin numbers don't match, this analysis is probably wrong!")
                bin_yields = []
                bin_codes = []
                for i,(name,ds) in enumerate(lib.items()):
                    missing_bins = [[b, ds.TestDefinition.getNameForBinno(b), ds.TestDefinition.getPlanForBinno(b), (ds.Data["Bin#"] == b).sum()] for b in list(set(ds.TestDefinition.getBinList()))]
                    if i == 0:
                        bin_yields = pd.DataFrame(missing_bins, columns=["Bin Code", "Alias", "Plan", f"{ds.Label}"]).set_index("Alias")
                    else:
                        bin_yields[f"{ds.Label}"] = pd.DataFrame(missing_bins, columns=["Bin Code", "Alias", "Plan", f"{ds.Label}"]).set_index("Alias")[f"{ds.Label}"]
                count_cols = [f"{ds.Label}" for ds in lib.values()]
                bin_yields["Total Count"] = bin_yields[count_cols].sum(axis=1)
                bin_yields = bin_yields.reset_index(names='Alias')
                bin_yields_styled = bin_yields.style.apply(highlight_increase, axis=1, subset=count_cols[1:]).map(color_zero, subset="Total Count")
                driver_bin_selection = st.dataframe(bin_yields_styled, hide_index=False, row_height=20, height="content", on_select="rerun", selection_mode="single-row")

                if driver_bin_selection.selection.rows:
                    show_bin_details(bin_yields.iloc[[driver_bin_selection.selection.rows[0]]], cmap)
                


        elif eng_task == "Find irregular bins":
            st.warning("Not implemented yet. Please check back later.")

        
        

